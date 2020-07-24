import networkx as nx
import timeit
import math
import matplotlib.pyplot as plt
import EmbeddingAndKmeans
import EmbeddingAndKmeans2
import random
import snap

import pandas as pd
import numpy as np
from openpyxl import load_workbook


def topk():
    compress = 0.1
    # t%
    t = 10
    count = 50
    print("start top-k query")
    start = timeit.default_timer()
    for i in range(count):
        G = nx.read_edgelist('Database/email-Enron.txt')
        pagerank = nx.pagerank(G)
        pagerank = sorted(pagerank.items(), key=lambda x: x[1], reverse=True)  # list
        nodes_G = []
        k = round(len(nx.nodes(G)) * t / 100)
        for i in range(k):
            key = pagerank[i][0]
            nodes_G.append(key)
    end = timeit.default_timer()
    with open("topk/email.txt", 'a') as file:
        string = "initial--- compression:" + str(compress) + " time:" + str((end - start) / count) + '\n'
        file.write(string)

    start = timeit.default_timer()
    for i in range(count):
        Gs1 = nx.read_edgelist("Data/minEmail-edges-0.1.txt")
        isolatednodes = getIsolatedNodes("Data/minEmail-edges-0.1.txt", "Data/minEmail-nodes-0.1.txt")
        Gs1.add_nodes_from(isolatednodes)
        k1 = round(len(nx.nodes(Gs1)) * t / 100)
        pagerank = nx.pagerank(Gs1)
        pagerank = sorted(pagerank.items(), key=lambda x: x[1], reverse=True)  # 变成list了
        nodes_GS1 = []
        for i in range(k1):
            key = pagerank[i][0]
            nodes_GS1.append(key)
    end = timeit.default_timer()
    with open("topk/email.txt", 'a') as file:
        string = "UDS--- compression:" + str(compress) + " time:" + str((end - start) / count) + '\n'
        file.write(string)

    start = timeit.default_timer()
    for i in range(count):
        Gs2 = nx.read_edgelist("Data/minEmail-edges-ADR-0.1.txt")
        k2 = round(len(nx.nodes(Gs2)) * t / 100)
        pagerank = nx.pagerank(Gs2)
        pagerank = sorted(pagerank.items(), key=lambda x: x[1], reverse=True)  # 变成list了
        nodes_GS2 = []
        for i in range(k):
            key = pagerank[i][0]
            nodes_GS2.append(key)
    end = timeit.default_timer()
    with open("topk/email.txt", 'a') as file:
        string = "ADR--- compression:" + str(compress) + " time:" + str((end - start) / count) + '\n'
        file.write(string)

    start = timeit.default_timer()
    for i in range(count):
        Gs3 = nx.read_edgelist("Data/minEmail-edges-ABM-0.1.txt")
        k3 = round(len(nx.nodes(Gs3)) * t / 100)
        pagerank = nx.pagerank(Gs3)
        pagerank = sorted(pagerank.items(), key=lambda x: x[1], reverse=True)  # 变成list了
        nodes_GS3 = []
        for i in range(k):
            key = pagerank[i][0]
            nodes_GS3.append(key)
    end = timeit.default_timer()
    with open("topk/email.txt", 'a') as file:
        string = "ABM--- compression:" + str(compress) + " time:" + str((end - start) / count) + '\n'
        file.write(string)

    # cal utiltiy
    utility = 0
    for supernode in nodes_GS1:
        node = []
        if ',' in supernode:  # supernode
            node = supernode.split(',')
        else:  # single node
            node.append(supernode)
        for temp in node:
            if temp in nodes_G:
                utility += 1 / len(node)
    utility /= k
    with open("topk/email.txt", 'a') as file:
        string = " UDS-utility:" + str(utility) + '\n'
        file.write(string)

    utility = 0
    for node in nodes_GS2:
        if node in nodes_G:
            utility += 1
    utility /= k
    with open("topk/email.txt", 'a') as file:
        string = " ADR-utility:" + str(utility) + '\n'
        file.write(string)

    utility = 0
    for node in nodes_GS3:
        if node in nodes_G:
            utility += 1
    utility /= k
    with open("topk/email.txt", 'a') as file:
        string = " ABM-utility:" + str(utility) + '\n'
        file.write(string)


def link():
    # 进行测试
    compress = 0.6
    count = 5

    print("start link prediction")
    # generate model by node2vec, use it by K-means

    # start = timeit.default_timer()
    # path = "Database/email-Enron.txt"
    # store = "embedding/email.emb"
    # store00 = "embedding/email.model"
    # # G = nx.read_edgelist(path)
    # EmbeddingAndKmeans.getModel(path, store, store00)
    # end = timeit.default_timer()
    # s = timeit.default_timer()
    # for i in range(count):
    #     node2label_G = EmbeddingAndKmeans.doCluster(store00)
    # e = timeit.default_timer()
    # with open("link/email.txt", 'a') as file:
    #     string = "initial---  compression:" + str(compress) + " time:" + str(end - start + (e - s) / count) + '\n'
    #     file.write(string)
    #
    # start = timeit.default_timer()
    # path1 = "Data/minEmail-edges-0.6.txt"
    # store1 = "embedding/minEmail-0.6.emb"
    # store11 = "embedding/minEmail-0.6.model"
    # Gs1 = nx.read_edgelist(path1)
    # isolatednodes = getIsolatedNodes(path1, "Data/minEmail-nodes-0.6.txt")
    # Gs1.add_nodes_from(isolatednodes)
    # EmbeddingAndKmeans2.getModel(Gs1, store1, store11)
    # end = timeit.default_timer()
    # s = timeit.default_timer()
    # for i in range(count):
    #     node2label_Gs1 = EmbeddingAndKmeans2.doCluster(store11)
    # e = timeit.default_timer()
    # with open("link/email.txt", 'a') as file:
    #     string = "UDS---  compression:" + str(compress) + " time:" + str(end - start + (e - s) / count) + '\n'
    #     file.write(string)
    #
    # start = timeit.default_timer()
    # path2 = "Data/minEmail-edges-ADR-0.1.txt"
    # store2 = "embedding/minEmail-ADR--0.1.emb"
    # store22 = "embedding/minEmail-ADR-0.1.model"
    # EmbeddingAndKmeans.getModel(path2, store2, store22)
    # end = timeit.default_timer()
    # s = timeit.default_timer()
    # # Gs2 = nx.read_edgelist(path2)
    # for i in range(count):
    #     node2label_Gs2 = EmbeddingAndKmeans.doCluster(store22)
    # e = timeit.default_timer()
    # with open("link/email.txt", 'a') as file:
    #     string = "ADR---  compression:" + str(compress) + " time:" + str(end - start + (e - s) / count) + '\n'
    #     # string = "ADR---  compression:" + str(compress) + " time1:" + str(end - start) + " time2:" + str(e - s) + '\n'
    #     file.write(string)
    #
    # start = timeit.default_timer()
    # path3 = "Data/minEmail-edges-ABM-0.1.txt"
    # store3 = "embedding/minEmail-ABM-0.1.emb"
    # store33 = "embedding/minEmail-ABM-0.1.model"
    # EmbeddingAndKmeans.getModel(path3, store3, store33)
    # end = timeit.default_timer()
    # s = timeit.default_timer()
    # for i in range(count):
    #     Gs3 = nx.read_edgelist(path3)
    #     node2label_Gs3 = EmbeddingAndKmeans.doCluster(store33)
    # e = timeit.default_timer()
    # with open("link/email.txt", 'a') as file:
    #     string = "ABM---  compression:" + str(compress) + " time:" + str(end - start + (e - s) / count) + '\n'
    #     file.write(string)

    # get the utility
    path = "Database/email-Enron.txt"
    path2 = "Data/minEmail-edges-ADR-0.6.txt"
    path3 = "Data/minEmail-edges-ABM-0.6.txt"
    store00 = "embedding/email.model"
    store11 = "embedding/minEmail-0.6.model"
    store22 = "embedding/minEmail-ADR-0.6.model"
    store33 = "embedding/minEmail-ABM-0.6.model"

    G = nx.read_edgelist(path)
    Gs2 = nx.read_edgelist(path2)
    Gs3 = nx.read_edgelist(path3)

    nodes1 = nx.nodes(G)
    nodelist1 = []
    for node in nodes1:
        nodelist1.append(node)
    nodes2 = nx.nodes(Gs2)
    nodelist2 = []
    for node in nodes2:
        nodelist2.append(node)
    nodes3 = nx.nodes(Gs3)
    nodelist3 = []
    for node in nodes3:
        nodelist3.append(node)

    total = 100
    totalnumber1 = 0
    totalnumber2 = 0
    totalnumber3 = 0
    # UDS
    for i in range(count):
        node2label_G = EmbeddingAndKmeans.doCluster(store00)
        node2label_Gs1 = EmbeddingAndKmeans2.doCluster(store11)
        number = 0
        for i in range(int(total / 2)):
            a = random.choice(nodelist1)
            b = a
            while a == b:
                b = random.choice(nodelist1)
            type1 = node2label_G[a]
            type2 = node2label_G[b]
            type3 = node2label_Gs1[a]
            type4 = node2label_Gs1[b]
            same1 = (type1 == type2)
            same2 = (type3 == type4)
            if same1 == same2:
                number += 1
        number2 = 0
        for i in range(int(total / 2)):
            a = random.choice(nodelist1)
            type = node2label_Gs1[a]
            # 找到与a相同聚类的点
            candidates = []
            for node in node2label_Gs1:
                if node2label_Gs1[node] == type:
                    candidates.append(node)
            b = a
            while a == b:
                b = random.choice(candidates)
            if node2label_G[a] == node2label_G[b]:
                number2 += 1
        totalnumber1 += number + number2
    # ADR
    for i in range(count):
        node2label_G = EmbeddingAndKmeans.doCluster(store00)
        node2label_Gs2 = EmbeddingAndKmeans.doCluster(store22)
        number = 0
        for i in range(int(total / 2)):
            a = random.choice(nodelist2)
            b = a
            while a == b:
                b = random.choice(nodelist2)
            type1 = node2label_G[a]
            type2 = node2label_G[b]
            type3 = node2label_Gs2[a]
            type4 = node2label_Gs2[b]
            same1 = (type1 == type2)
            same2 = (type3 == type4)
            if same1 == same2:
                number += 1
        number2 = 0
        for i in range(int(total / 2)):
            a = random.choice(nodelist2)
            type = node2label_Gs2[a]
            # 找到与a相同聚类的点
            candidates = []
            for node in node2label_Gs2:
                if node2label_Gs2[node] == type:
                    candidates.append(node)
            b = a
            while a == b:
                b = random.choice(candidates)
            if node2label_G[a] == node2label_G[b]:
                number2 += 1
        totalnumber2 += number + number2
    # ABM
    for i in range(count):
        node2label_G = EmbeddingAndKmeans.doCluster(store00)
        node2label_Gs3 = EmbeddingAndKmeans.doCluster(store33)
        number = 0
        for i in range(int(total / 2)):
            a = random.choice(nodelist3)
            b = a
            while a == b:
                b = random.choice(nodelist3)
            type1 = node2label_G[a]
            type2 = node2label_G[b]
            type3 = node2label_Gs3[a]
            type4 = node2label_Gs3[b]
            same1 = (type1 == type2)
            same2 = (type3 == type4)
            if same1 == same2:
                number += 1
        number2 = 0
        for i in range(int(total / 2)):
            a = random.choice(nodelist3)
            type = node2label_Gs3[a]
            # 找到与a相同聚类的点
            candidates = []
            for node in node2label_Gs3:
                if node2label_Gs3[node] == type:
                    candidates.append(node)
            b = a
            while a == b:
                b = random.choice(candidates)
            if node2label_G[a] == node2label_G[b]:
                number2 += 1
        totalnumber3 += number + number2
    utility1 = totalnumber1 / (total * count)
    utility2 = totalnumber2 / (total * count)
    utility3 = totalnumber3 / (total * count)
    with open("link/email-link.txt", 'a') as file:
        string1 = " UDS-utility:" + str(utility1) + " compression:" + str(compress) + " pairs:" + str(total) + '\n'
        string2 = " ADR-utility:" + str(utility2) + " compression:" + str(compress) + " pairs:" + str(total) + '\n'
        string3 = " ABM-utility:" + str(utility3) + " compression:" + str(compress) + " pairs:" + str(total) + '\n'
        file.write(string1)
        file.write(string2)
        file.write(string3)


def getIsolatedNodes(edgepath, nodepath):
    g = nx.read_edgelist(edgepath)
    list2 = []
    nodes1 = nx.nodes(g)
    for node in nodes1:
        list2.append(node)
    list1 = []
    with open(nodepath, 'r') as file:
        for line in file.readlines():
            line = line.strip('\n')
            list1.append(line)
    list3 = list(set(list1) - set(list2))
    return list3


def VertexDegree():
    compress = 0.1
    c = 10
    interval = 30
    start = timeit.default_timer()
    commontext = "Email"
    for i in range(c):
        G = nx.read_edgelist('Database/email-Enron.txt')
        degree = nx.degree_histogram(G)  # degree distribution
        degree_list = {}
        count = 0
        # i = 0
        for i in range(len(degree)):  # i->degree, degree[i]->nums
            count += degree[i]
            if i > 300:
                j = 299
            else:
                j = i
            key = int(math.floor(j) / interval) * interval
            # key = int(math.floor(i) / interval) * interval
            if key in degree_list:
                number = degree_list[key]
                degree_list[key] = number + degree[i]
            else:
                degree_list[key] = degree[i]
    end = timeit.default_timer()
    with open("degree/email.txt", 'a') as file:
        string = "initial---  "+"compression:" + str(compress) + " time:" + str((end - start) / c) + '\n'
        file.write(string)

    start = timeit.default_timer()
    for i in range(c):
        edgename = "Data/min" + commontext + "-edges-" + str(compress) + ".txt"
        nodename = "Data/min" + commontext + "-nodes-" + str(compress) + ".txt"
        Gs1 = nx.read_edgelist(edgename)
        isolatednodes = getIsolatedNodes(edgename, nodename)
        Gs1.add_nodes_from(isolatednodes)
        degree1 = nx.degree_histogram(Gs1)
        degree_list1 = {}
        count = 0
        i = 0
        for i in range(len(degree1)):
            count += degree1[i]
            if i > 300:
                j = 299
            else:
                j = i
            key = int(math.floor(j) / interval) * interval
            # key = int(math.floor(i) / interval) * interval
            if key in degree_list1:
                number = degree_list1[key]
                degree_list1[key] = number + degree1[i]
            else:
                degree_list1[key] = degree1[i]
    end = timeit.default_timer()
    with open("degree/email.txt", 'a') as file:
        string = "UDS---  " + "compression:" + str(compress) + " time:" + str((end - start) / c) + '\n'
        file.write(string)

    start = timeit.default_timer()
    for i in range(c):
        edgename = "Data/min" + commontext + "-edges-ADR-" + str(compress) + ".txt"
        Gs2 = nx.read_edgelist(edgename)
        degree2 = nx.degree_histogram(Gs2)
        degree_list2 = {}
        count = 0
        for i in range(len(degree2)):  # estimation
            count += degree2[i]
            j = i * (1 / compress)
            if j > 300:
                j = 299
            key = int(math.floor(j) / interval) * interval
            if key in degree_list2:
                number = degree_list2[key]
                degree_list2[key] = number + degree2[i]
            else:
                degree_list2[key] = degree2[i]
    end = timeit.default_timer()
    with open("degree/email.txt", 'a') as file:
        string = "ADR---  " + "compression:" + str(compress) + " time:" + str((end - start) / c) + '\n'
        file.write(string)

    start = timeit.default_timer()
    for i in range(c):
        edgename = "Data/min" + commontext + "-edges-ABM-" + str(compress) + ".txt"
        Gs3= nx.read_edgelist(edgename)
        degree3 = nx.degree_histogram(Gs3)
        degree_list3 = {}
        count = 0
        for i in range(len(degree3)):  # estimation
            count += degree3[i]
            j = i * (1 / compress)
            if j > 300:
                j = 299
            key = int(math.floor(j) / interval) * interval
            if key in degree_list3:
                number = degree_list3[key]
                degree_list3[key] = number + degree3[i]
            else:
                degree_list3[key] = degree3[i]
    end = timeit.default_timer()
    with open("degree/email.txt", 'a') as file:
        string = "ABM---  " + "compression:" + str(compress) + " time:" + str((end - start) / c) + '\n' + '\n'
        file.write(string)

    x = list(degree_list.keys())
    y = [z / float(sum(degree_list.values())) for z in degree_list.values()]  # initial
    x1 = list(degree_list1.keys())
    y1 = [z / float(sum(degree_list1.values())) for z in degree_list1.values()]  # UDS
    x2 = list(degree_list2.keys())
    y2 = [z / float(sum(degree_list2.values())) for z in degree_list2.values()]  # ADR
    x3 = list(degree_list3.keys())
    y3 = [z / float(sum(degree_list3.values())) for z in degree_list3.values()]  # ABM

    # save the data
    A = np.array([x,y])
    A1 = np.array([x1,y1])
    A2 = np.array([x2, y2])
    A3 = np.array([x3, y3])
    data = pd.DataFrame(A)
    data1 = pd.DataFrame(A1)
    data2 = pd.DataFrame(A2)
    data3 = pd.DataFrame(A3)

    book = load_workbook('Picture/degree.xlsx')
    writer = pd.ExcelWriter('Picture/degree.xlsx', engine='openpyxl')
    writer.book = book

    pagename = "email" + str(compress)
    data.to_excel(writer, sheet_name=pagename,float_format='%.5f')
    data1.to_excel(writer, sheet_name=pagename,startrow=3, float_format='%.5f')
    data2.to_excel(writer, sheet_name=pagename, startrow=6, float_format='%.5f')
    data3.to_excel(writer, sheet_name=pagename, startrow=9, float_format='%.5f')
    writer.save()

    writer.close()

    plt.plot(x, y, marker='o', mec='k', mfc='w', c='k', label=u'initial')
    plt.plot(x1, y1, marker='v', mec='#EEB422', mfc='w', c='#EEB422', label=u'UDS')
    plt.plot(x2, y2, marker='s', mec='r', mfc='w', c='r', label=u'ADR')
    plt.plot(x3, y3, marker='+', mec='#00BFFF', mfc='w', c='#00BFFF', label=u'ABM')

    plt.legend()
    plt.xticks(x, list(degree_list.keys()))
    plt.margins(0)
    plt.subplots_adjust(bottom=0.15)
    plt.xlabel(u"degree")  # X
    plt.ylabel("% of vertices")  # Y
    plt.title("email-Enron")  # title

    plt.show()


def ShortPath():
    compress = 0.9
    start = timeit.default_timer()
    G = nx.read_edgelist('Database/email-Enron.txt')
    commontext = "Email"
    sp = dict(nx.shortest_path_length(G))
    length2num = {}
    num = 0  # nums of reachable nodes
    for source in sp:
        num += (len(sp[source]) - 1)
        for target in sp[source]:
            length = sp[source][target]
            if length not in length2num.keys():
                length2num[length] = 1
            else:
                temp = length2num[length]
                length2num[length] = temp + 1
    length2num[0] = 0
    end = timeit.default_timer()
    with open("sp/email.txt", 'a') as file:
        string = "initial--- compression:" + str(compress) + " time:" + str(end - start) + '\n'
        file.write(string)

    start = timeit.default_timer()
    edgename = "Data/min" + commontext + "-edges-" + str(compress) + ".txt"
    nodename = "Data/min" + commontext + "-nodes-" + str(compress) + ".txt"
    Gs1 = nx.read_edgelist(edgename)
    isolatednodes = getIsolatedNodes(edgename, nodename)
    Gs1.add_nodes_from(isolatednodes)
    sp1 = dict(nx.shortest_path_length(Gs1))
    length2num1 = {}
    num1 = 0
    for source in sp1:
        num1 += (len(sp1[source]) - 1)
        for target in sp1[source]:
            length = round(sp1[source][target] * (compress + (1 - compress) * 0.5))
            if length not in length2num1.keys():
                length2num1[length] = 1
            else:
                temp = length2num1[length]
                length2num1[length] = temp + 1
    length2num1[0] = 0
    end = timeit.default_timer()
    with open("sp/email.txt", 'a') as file:
        string = "UDS--- compression:" + str(compress) + " time:" + str(end - start) + '\n'
        file.write(string)

    start = timeit.default_timer()
    edgename = "Data/min" + commontext + "-edges-ADR-" + str(compress) + ".txt"
    Gs2 = nx.read_edgelist(edgename)
    sp2 = dict(nx.shortest_path_length(Gs2))
    length2num2 = {}
    num2 = 0
    for source in sp2:
        num2 += (len(sp2[source]) - 1)
        for target in sp2[source]:
            length = round(sp2[source][target] * (compress + (1 - compress) * 0.5))

            if length not in length2num2.keys():
                length2num2[length] = 1
            else:
                temp = length2num2[length]
                length2num2[length] = temp + 1
    length2num2[0] = 0
    end = timeit.default_timer()
    with open("sp/email.txt", 'a') as file:
        string = "ADR--- compression:" + str(compress) + " time:" + str(end - start) + '\n'
        file.write(string)

    start = timeit.default_timer()
    edgename = "Data/min" + commontext + "-edges-ABM-" + str(compress) + ".txt"
    Gs3 = nx.read_edgelist(edgename)
    sp3 = dict(nx.shortest_path_length(Gs3))
    length2num3 = {}
    num3 = 0
    for source in sp3:
        num3 += (len(sp3[source]) - 1)
        for target in sp3[source]:
            length = round(sp3[source][target] * (compress + (1 - compress) * 0.5))

            if length not in length2num3.keys():
                length2num3[length] = 1
            else:
                temp = length2num3[length]
                length2num3[length] = temp + 1
    length2num3[0] = 0
    end = timeit.default_timer()
    with open("sp/email.txt", 'a') as file:
        string = "ABM--- compression:" + str(compress) + " time:" + str(end - start) + '\n'  + '\n'
        file.write(string)

    length2num = sorted(length2num.items(), key=lambda x: x[0])
    length2num1 = sorted(length2num1.items(), key=lambda x: x[0])
    length2num2 = sorted(length2num2.items(), key=lambda x: x[0])
    length2num3 = sorted(length2num3.items(), key=lambda x: x[0])

    x = [z[0] for z in length2num]
    y = [z[1] / num for z in length2num]
    x1 = [z[0] for z in length2num1]
    y1 = [z[1] / num1 for z in length2num1]
    x2 = [z[0] for z in length2num2]
    y2 = [z[1] / num2 for z in length2num2]
    x3 = [z[0] for z in length2num3]
    y3 = [z[1] / num3 for z in length2num3]

    # save
    A = np.array([x, y])
    A1 = np.array([x1, y1])
    A2 = np.array([x2, y2])
    A3 = np.array([x3, y3])

    data = pd.DataFrame(A)
    data1 = pd.DataFrame(A1)
    data2 = pd.DataFrame(A2)
    data3 = pd.DataFrame(A3)

    book = load_workbook('Picture/sp.xlsx')
    writer = pd.ExcelWriter('Picture/sp.xlsx', engine='openpyxl')
    writer.book = book

    pagename = "email" + str(compress)
    data.to_excel(writer, sheet_name=pagename, float_format='%.5f')
    data1.to_excel(writer, sheet_name=pagename, startrow=3, float_format='%.5f')
    data2.to_excel(writer, sheet_name=pagename, startrow=6, float_format='%.5f')
    data3.to_excel(writer, sheet_name=pagename, startrow=9, float_format='%.5f')
    writer.save()

    writer.close()

    plt.plot(x, y, marker='o', mec='k', mfc='w', c='k', label=u'initial')
    plt.plot(x1, y1, marker='v', mec='#EEB422', mfc='w', c='#EEB422', label=u'UDS')
    plt.plot(x2, y2, marker='s', mec='r', mfc='w', c='r', label=u'ADR')
    plt.plot(x3, y3, marker='+', mec='#00BFFF', mfc='w', c='#00BFFF', label=u'ABM')

    plt.legend()
    plt.xticks(x, [z[0] for z in length2num])
    plt.margins(0)
    plt.subplots_adjust(bottom=0.15)
    plt.xlabel(u"SP distance")  # X
    plt.ylabel("% of vertex pairs")  # Y
    plt.title("email-Enron")  # title

    plt.show()


def centrality():
    compress = 0.6
    interval = 50
    count = 1
    commontext = "HepPh"
    start = timeit.default_timer()
    for i in range(count):
        G = nx.read_edgelist('Database/ca-HepPh.txt')
        degree2bc = {}
        degree2count = {}  # nums
        bc = nx.betweenness_centrality(G)

        degree = nx.degree(G)
        for tuple in degree:
            node = tuple[0]
            d = tuple[1]

            key = int(math.floor(d / interval)) * interval
            if key not in degree2bc.keys():
                degree2bc[key] = bc[node]
                degree2count[key] = 1
            else:
                temp = degree2bc[key]
                degree2bc[key] = temp + bc[node]
                tempcount = degree2count[key]
                degree2count[key] = tempcount + 1
        for d in degree2bc:
            centrality = degree2bc[d]
            degree2bc[d] = centrality / degree2count[d]
    end = timeit.default_timer()
    with open("centrality/caHepPh.txt", 'a') as file:
        string = "initial--- compression:" + str(compress) + " time:" + str((end - start) / count) + '\n'
        file.write(string)

    start = timeit.default_timer()
    for i in range(count):
        edgename = "Data/min" + commontext + "-edges-" + str(compress) + ".txt"
        nodename = "Data/min" + commontext + "-nodes-" + str(compress) + ".txt"
        Gs1 = nx.read_edgelist(edgename)
        isolatednodes = getIsolatedNodes(edgename, nodename)
        Gs1.add_nodes_from(isolatednodes)
        degree2bc1 = {}
        degree2count1 = {}
        bc1 = nx.betweenness_centrality(Gs1)

        degree1 = nx.degree(Gs1)
        for tuple in degree1:
            node = tuple[0]
            d = tuple[1]

            key = int(math.floor(d / interval)) * interval
            if key not in degree2bc1.keys():
                degree2bc1[key] = bc1[node]
                degree2count1[key] = 1
            else:
                temp = degree2bc1[key]
                degree2bc1[key] = temp + bc1[node]
                tempcount = degree2count1[key]
                degree2count1[key] = tempcount + 1
        for d in degree2bc1:
            centrality = degree2bc1[d]
            degree2bc1[d] = centrality / degree2count1[d]
    end = timeit.default_timer()
    with open("centrality/caHepPh.txt", 'a') as file:
        string = "UDS--- compression:" + str(compress) + " time:" + str((end - start) / count) + '\n'
        file.write(string)

    start = timeit.default_timer()
    for i in range(count):
        edgename = "Data/min" + commontext + "-edges-ADR-" + str(compress) + ".txt"
        Gs2 = nx.read_edgelist(edgename)
        degree2bc2 = {}
        degree2count2 = {}
        bc2 = nx.betweenness_centrality(Gs2)

        degree2 = nx.degree(Gs2)
        for tuple in degree2:
            node = tuple[0]
            d = tuple[1]

            key = int(math.floor((d / compress) / interval)) * interval
            if key not in degree2bc2.keys():
                degree2bc2[key] = bc2[node]  # / ((2 - compress) * compress)
                degree2count2[key] = 1
            else:
                temp = degree2bc2[key]
                degree2bc2[key] = temp + bc2[node]  # / ((2 - compress) * compress)
                tempcount = degree2count2[key]
                degree2count2[key] = tempcount + 1
        for d in degree2bc2:
            centrality = degree2bc2[d]
            degree2bc2[d] = centrality / degree2count2[d]
    end = timeit.default_timer()
    with open("centrality/caHepPh", 'a') as file:
        string = "ADR--- compression:" + str(compress) + " time:" + str((end - start) / count) + '\n'
        file.write(string)

    start = timeit.default_timer()
    for i in range(count):
        edgename = "Data/min" + commontext + "-edges-ABM-" + str(compress) + ".txt"
        Gs3 = nx.read_edgelist(edgename)
        degree2bc3 = {}
        degree2count3 = {}
        bc3 = nx.betweenness_centrality(Gs3)

        degree3 = nx.degree(Gs3)
        for tuple in degree3:
            node = tuple[0]
            d = tuple[1]

            key = int(math.floor((d / compress) / interval)) * interval
            if key not in degree2bc3.keys():
                degree2bc3[key] = bc3[node]  # / ((2 - compress) * compress)
                degree2count3[key] = 1
            else:
                temp = degree2bc3[key]
                degree2bc3[key] = temp + bc3[node]  # / ((2 - compress) * compress)
                tempcount = degree2count3[key]
                degree2count3[key] = tempcount + 1
        for d in degree2bc3:
            centrality = degree2bc3[d]
            degree2bc3[d] = centrality / degree2count3[d]
    end = timeit.default_timer()
    with open("centrality/caHepPh.txt", 'a') as file:
        string = "ABM--- compression:" + str(compress) + " time:" + str((end - start) / count) + '\n' + '\n'
        file.write(string)

    degree2bc = sorted(degree2bc.items(), key=lambda x: x[0])
    degree2bc1 = sorted(degree2bc1.items(), key=lambda x: x[0])
    degree2bc2 = sorted(degree2bc2.items(), key=lambda x: x[0])
    degree2bc3 = sorted(degree2bc3.items(), key=lambda x: x[0])

    x = [z[0] for z in degree2bc]
    y = [z[1] for z in degree2bc]
    x1 = [z[0] for z in degree2bc1]
    y1 = [z[1] for z in degree2bc1]
    x2 = [z[0] for z in degree2bc2]
    y2 = [z[1] for z in degree2bc2]
    x3 = [z[0] for z in degree2bc3]
    y3 = [z[1] for z in degree2bc3]

    # save
    A = np.array([x, y])
    A1 = np.array([x1, y1])
    A2 = np.array([x2, y2])
    A3 = np.array([x3, y3])

    data = pd.DataFrame(A)
    data1 = pd.DataFrame(A1)
    data2 = pd.DataFrame(A2)
    data3 = pd.DataFrame(A3)

    book = load_workbook('Picture/bc.xlsx')
    writer = pd.ExcelWriter('Picture/bc.xlsx', engine='openpyxl')
    writer.book = book

    pagename = "caHepPh" + str(compress)
    data.to_excel(writer, sheet_name=pagename, float_format='%.5f')
    data1.to_excel(writer, sheet_name=pagename, startrow=3, float_format='%.5f')
    data2.to_excel(writer, sheet_name=pagename, startrow=6, float_format='%.5f')
    data3.to_excel(writer, sheet_name=pagename, startrow=9, float_format='%.5f')
    writer.save()

    writer.close()

    plt.plot(x, y, marker='o', mec='k', mfc='w', c='k', label=u'initial')
    plt.plot(x1, y1, marker='v', mec='#EEB422', mfc='w', c='#EEB422', label=u'UDS')
    plt.plot(x2, y2, marker='s', mec='r', mfc='w', c='r', label=u'ADR')
    plt.plot(x3, y3, marker='+', mec='#00BFFF', mfc='w', c='#00BFFF', label=u'ABM')

    plt.legend()
    plt.xticks(x, [z[0] for z in degree2bc])
    plt.margins(0)
    plt.subplots_adjust(bottom=0.15)
    plt.xlabel(u"degree")  # X
    plt.ylabel(u"betweenness centrality")  # Y
    plt.title("ca-HepPh")  # title

    plt.show()


def coefficient():
    compress = 0.9
    commontext = "GrQc"
    interval = 10
    start = timeit.default_timer()
    G = nx.read_edgelist('Database/ca-GrQc.txt')
    degree2cc = {}
    degree2ccount = {}
    cc = nx.clustering(G)
    degree = nx.degree(G)
    for tuple in degree:
        node = tuple[0]
        d = tuple[1]

        key = int(math.floor(d / interval)) * interval
        if key not in degree2cc.keys():
            degree2cc[key] = cc[node]
            degree2ccount[key] = 1
        else:
            temp = degree2cc[key]
            degree2cc[key] = temp + cc[node]
            tempcount = degree2ccount[key]
            degree2ccount[key] = tempcount + 1

    for c in degree2cc:
        coeff = degree2cc[c]
        degree2cc[c] = coeff / degree2ccount[c]
    end = timeit.default_timer()
    # del G
    # del cc
    # del degree
    # del degree2ccount
    with open("cc/caGrQc.txt", 'a') as file:
        string = "initial--- compression:" + str(compress) + " time:" + str(end - start) + '\n'
        file.write(string)

    start = timeit.default_timer()
    edgename = "Data/min" + commontext + "-edges-" + str(compress) + ".txt"
    nodename = "Data/min" + commontext + "-nodes-" + str(compress) + ".txt"
    Gs1 = nx.read_edgelist(edgename)
    isolatednodes = getIsolatedNodes(edgename, nodename)
    Gs1.add_nodes_from(isolatednodes)
    degree2cc1 = {}
    degree2ccount1 = {}  # nums
    cc1 = nx.clustering(Gs1)
    degree1 = nx.degree(Gs1)
    for tuple in degree1:
        node = tuple[0]
        d = tuple[1]

        key = int(math.floor(d / interval)) * interval
        if key not in degree2cc1.keys():
            degree2cc1[key] = cc1[node]
            degree2ccount1[key] = 1
        else:
            temp = degree2cc1[key]
            degree2cc1[key] = temp + cc1[node]
            tempcount = degree2ccount1[key]
            degree2ccount1[key] = tempcount + 1
    for d in degree2cc1:
        coeff = degree2cc1[d]
        degree2cc1[d] = coeff / degree2ccount1[d]
    end = timeit.default_timer()
    # del Gs1
    # del cc1
    # del degree1
    # del degree2ccount1
    with open("cc/caGrQc.txt", 'a') as file:
        string = "UDS--- compression:" + str(compress) + " time:" + str(end - start) + '\n'
        file.write(string)

    start = timeit.default_timer()
    edgename = "Data/min" + commontext + "-edges-ADR-" + str(compress) + ".txt"
    Gs2 = nx.read_edgelist(edgename)
    degree2cc2 = {}
    degree2ccount2 = {}  # nums
    cc2 = nx.clustering(Gs2)
    degree2 = nx.degree(Gs2)
    for tuple in degree2:
        node = tuple[0]
        d = tuple[1]

        key = int(math.floor((d / compress) / interval)) * interval
        if key not in degree2cc2.keys():
            degree2cc2[key] = cc2[node] / compress
            degree2ccount2[key] = 1
        else:
            temp = degree2cc2[key]
            degree2cc2[key] = temp + cc2[node] / compress
            tempcount = degree2ccount2[key]
            degree2ccount2[key] = tempcount + 1
    for d in degree2cc2:
        coeff = degree2cc2[d]
        degree2cc2[d] = coeff / degree2ccount2[d]
    end = timeit.default_timer()
    # del Gs2
    # del cc2
    # del degree2
    # del degree2ccount2
    with open("cc/caGrQc.txt", 'a') as file:
        string = "ADR--- compression:" + str(compress) + " time:" + str(end - start) + '\n'
        file.write(string)

    start = timeit.default_timer()
    edgename = "Data/min" + commontext + "-edges-ABM-" + str(compress) + ".txt"
    Gs3 = nx.read_edgelist(edgename)
    degree2cc3 = {}
    degree2ccount3 = {}  # nums
    cc3 = nx.clustering(Gs3)
    degree3 = nx.degree(Gs3)
    for tuple in degree3:
        node = tuple[0]
        d = tuple[1]

        key = int(math.floor((d / compress) / interval)) * interval
        if key not in degree2cc3.keys():
            degree2cc3[key] = cc3[node] / compress
            degree2ccount3[key] = 1
        else:
            temp = degree2cc3[key]
            degree2cc3[key] = temp + cc3[node] / compress
            tempcount = degree2ccount3[key]
            degree2ccount3[key] = tempcount + 1
    for d in degree2cc3:
        coeff = degree2cc3[d]
        degree2cc3[d] = coeff / degree2ccount3[d]
    end = timeit.default_timer()
    # del Gs3
    # del cc3
    # del degree3
    # del degree2ccount3
    with open("cc/caGrQc.txt", 'a') as file:
        string = "ABM--- compression:" + str(compress) + " time:" + str(end - start) + '\n'
        file.write(string)

    degree2cc = sorted(degree2cc.items(), key=lambda x: x[0])
    degree2cc1 = sorted(degree2cc1.items(), key=lambda x: x[0])
    degree2cc2 = sorted(degree2cc2.items(), key=lambda x: x[0])
    degree2cc3 = sorted(degree2cc3.items(), key=lambda x: x[0])

    x = [z[0] for z in degree2cc]
    y = [z[1] for z in degree2cc]
    x1 = [z[0] for z in degree2cc1]
    y1 = [z[1] for z in degree2cc1]
    x2 = [z[0] for z in degree2cc2]
    y2 = [z[1] for z in degree2cc2]
    x3 = [z[0] for z in degree2cc3]
    y3 = [z[1] for z in degree2cc3]

    # save
    A = np.array([x, y])
    A1 = np.array([x1, y1])
    A2 = np.array([x2, y2])
    A3 = np.array([x3, y3])

    data = pd.DataFrame(A)
    data1 = pd.DataFrame(A1)
    data2 = pd.DataFrame(A2)
    data3 = pd.DataFrame(A3)

    book = load_workbook('Picture/cc.xlsx')
    writer = pd.ExcelWriter('Picture/cc.xlsx', engine='openpyxl')
    writer.book = book

    pagename = "caGrQc" + str(compress)
    data.to_excel(writer, sheet_name=pagename, float_format='%.5f')
    data1.to_excel(writer, sheet_name=pagename, startrow=3, float_format='%.5f')
    data2.to_excel(writer, sheet_name=pagename, startrow=6, float_format='%.5f')
    data3.to_excel(writer, sheet_name=pagename, startrow=9, float_format='%.5f')
    writer.save()

    writer.close()

    plt.plot(x, y, marker='o', mec='k', mfc='w', c='k', label=u'initial')
    plt.plot(x1, y1, marker='v', mec='#EEB422', mfc='w', c='#EEB422', label=u'UDS')
    plt.plot(x2, y2, marker='s', mec='r', mfc='w', c='r', label=u'ADR')
    plt.plot(x3, y3, marker='+', mec='#00BFFF', mfc='w', c='#00BFFF', label=u'ABM')

    plt.legend()
    plt.xticks(x, [z[0] for z in degree2cc])
    plt.margins(0)
    plt.subplots_adjust(bottom=0.15)
    plt.xlabel(u"degree")  # X
    plt.ylabel(u"clustering coefficient")  # Y
    plt.title("ca-GrQc")  # title

    plt.show()


def hopplot():
    compress = 0.1
    base = 2
    commontext = "GrQc"

    start = timeit.default_timer()
    G = nx.read_edgelist('Database/ca-GrQc.txt')
    paths = dict(nx.shortest_path_length(G))
    set = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    num = 0  # nums of all reachable nodes
    for source in paths:
        num += (len(paths[source]) - 1)
        for target in paths[source]:
            length = paths[source][target]
            if length < base:
                set[1] = set[1] + 1
                set[2] = set[2] + 1
                set[3] = set[3] + 1
                set[4] = set[4] + 1
                set[5] = set[5] + 1
                set[6] = set[6] + 1
                set[7] = set[7] + 1
                set[8] = set[8] + 1
                set[9] = set[9] + 1
                set[10] = set[10] + 1
            elif length < base * 2:
                set[2] = set[2] + 1
                set[3] = set[3] + 1
                set[4] = set[4] + 1
                set[5] = set[5] + 1
                set[6] = set[6] + 1
                set[7] = set[7] + 1
                set[8] = set[8] + 1
                set[9] = set[9] + 1
                set[10] = set[10] + 1
            elif length < base * 3:
                set[3] = set[3] + 1
                set[4] = set[4] + 1
                set[5] = set[5] + 1
                set[6] = set[6] + 1
                set[7] = set[7] + 1
                set[8] = set[8] + 1
                set[9] = set[9] + 1
                set[10] = set[10] + 1
            elif length < base * 4:
                set[4] = set[4] + 1
                set[5] = set[5] + 1
                set[6] = set[6] + 1
                set[7] = set[7] + 1
                set[8] = set[8] + 1
                set[9] = set[9] + 1
                set[10] = set[10] + 1
            elif length < base * 5:
                set[5] = set[5] + 1
                set[6] = set[6] + 1
                set[7] = set[7] + 1
                set[8] = set[8] + 1
                set[9] = set[9] + 1
                set[10] = set[10] + 1
            elif length < base * 6:
                set[6] = set[6] + 1
                set[7] = set[7] + 1
                set[8] = set[8] + 1
                set[9] = set[9] + 1
                set[10] = set[10] + 1
            elif length < base * 7:
                set[7] = set[7] + 1
                set[8] = set[8] + 1
                set[9] = set[9] + 1
                set[10] = set[10] + 1
            elif length < base * 8:
                set[8] = set[8] + 1
                set[9] = set[9] + 1
                set[10] = set[10] + 1
            elif length < base * 9:
                set[9] = set[9] + 1
                set[10] = set[10] + 1
            elif length < base * 10:
                set[10] = set[10] + 1
    end = timeit.default_timer()
    with open("hopplot/caGrQc.txt", 'a') as file:
        string = "initial--- compression:" + str(compress) + " time:" + str(end - start) + '\n'
        file.write(string)

    start = timeit.default_timer()
    edgename = "Data/min" + commontext + "-edges-" + str(compress) + ".txt"
    nodename = "Data/min" + commontext + "-nodes-" + str(compress) + ".txt"
    Gs1 = nx.read_edgelist(edgename)
    isolatednodes = getIsolatedNodes(edgename, nodename)
    Gs1.add_nodes_from(isolatednodes)
    paths1 = dict(nx.shortest_path_length(Gs1))
    set1 = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    num1 = 0
    for source in paths1:
        num1 += (len(paths1[source]) - 1)
        for target in paths1[source]:
            length = paths1[source][target]
            if length < base:
                set1[1] = set1[1] + 1
                set1[2] = set1[2] + 1
                set1[3] = set1[3] + 1
                set1[4] = set1[4] + 1
                set1[5] = set1[5] + 1
                set1[6] = set1[6] + 1
                set1[7] = set1[7] + 1
                set1[8] = set1[8] + 1
                set1[9] = set1[9] + 1
                set1[10] = set1[10] + 1
            elif length < base * 2:
                set1[2] = set1[2] + 1
                set1[3] = set1[3] + 1
                set1[4] = set1[4] + 1
                set1[5] = set1[5] + 1
                set1[6] = set1[6] + 1
                set1[7] = set1[7] + 1
                set1[8] = set1[8] + 1
                set1[9] = set1[9] + 1
                set1[10] = set1[10] + 1
            elif length < base * 3:
                set1[3] = set1[3] + 1
                set1[4] = set1[4] + 1
                set1[5] = set1[5] + 1
                set1[6] = set1[6] + 1
                set1[7] = set1[7] + 1
                set1[8] = set1[8] + 1
                set1[9] = set1[9] + 1
                set1[10] = set1[10] + 1
            elif length < base * 4:
                set1[4] = set1[4] + 1
                set1[5] = set1[5] + 1
                set1[6] = set1[6] + 1
                set1[7] = set1[7] + 1
                set1[8] = set1[8] + 1
                set1[9] = set1[9] + 1
                set1[10] = set1[10] + 1
            elif length < base * 5:
                set1[5] = set1[5] + 1
                set1[6] = set1[6] + 1
                set1[7] = set1[7] + 1
                set1[8] = set1[8] + 1
                set1[9] = set1[9] + 1
                set1[10] = set1[10] + 1
            elif length < base * 6:
                set1[6] = set1[6] + 1
                set1[7] = set1[7] + 1
                set1[8] = set1[8] + 1
                set1[9] = set1[9] + 1
                set1[10] = set1[10] + 1
            elif length < base * 7:
                set1[7] = set1[7] + 1
                set1[8] = set1[8] + 1
                set1[9] = set1[9] + 1
                set1[10] = set1[10] + 1
            elif length < base * 8:
                set1[8] = set1[8] + 1
                set1[9] = set1[9] + 1
                set1[10] = set1[10] + 1
            elif length < base * 9:
                set1[9] = set1[9] + 1
                set1[10] = set1[10] + 1
            elif length < base * 10:
                set1[10] = set1[10] + 1
    end = timeit.default_timer()
    with open("hopplot/caGrQc.txt", 'a') as file:
        string = "UDS--- compression:" + str(compress) + " time:" + str(end - start) + '\n'
        file.write(string)

    start = timeit.default_timer()
    edgename = "Data/min" + commontext + "-edges-ADR-" + str(compress) + ".txt"
    Gs2 = nx.read_edgelist(edgename)
    paths2 = dict(nx.shortest_path_length(Gs2))
    set2 = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    num2 = 0
    for source in paths2:
        num2 += (len(paths2[source]) - 1)
        for target in paths2[source]:
            length = paths2[source][target] *(compress+0.65*(1-compress))
            if length < base:
                set2[1] = set2[1] + 1
                set2[2] = set2[2] + 1
                set2[3] = set2[3] + 1
                set2[4] = set2[4] + 1
                set2[5] = set2[5] + 1
                set2[6] = set2[6] + 1
                set2[7] = set2[7] + 1
                set2[8] = set2[8] + 1
                set2[9] = set2[9] + 1
                set2[10] = set2[10] + 1
            elif length < base * 2:
                set2[2] = set2[2] + 1
                set2[3] = set2[3] + 1
                set2[4] = set2[4] + 1
                set2[5] = set2[5] + 1
                set2[6] = set2[6] + 1
                set2[7] = set2[7] + 1
                set2[8] = set2[8] + 1
                set2[9] = set2[9] + 1
                set2[10] = set2[10] + 1
            elif length < base * 3:
                set2[3] = set2[3] + 1
                set2[4] = set2[4] + 1
                set2[5] = set2[5] + 1
                set2[6] = set2[6] + 1
                set2[7] = set2[7] + 1
                set2[8] = set2[8] + 1
                set2[9] = set2[9] + 1
                set2[10] = set2[10] + 1
            elif length < base * 4:
                set2[4] = set2[4] + 1
                set2[5] = set2[5] + 1
                set2[6] = set2[6] + 1
                set2[7] = set2[7] + 1
                set2[8] = set2[8] + 1
                set2[9] = set2[9] + 1
                set2[10] = set2[10] + 1
            elif length < base * 5:
                set2[5] = set2[5] + 1
                set2[6] = set2[6] + 1
                set2[7] = set2[7] + 1
                set2[8] = set2[8] + 1
                set2[9] = set2[9] + 1
                set2[10] = set2[10] + 1
            elif length < base * 6:
                set2[6] = set2[6] + 1
                set2[7] = set2[7] + 1
                set2[8] = set2[8] + 1
                set2[9] = set2[9] + 1
                set2[10] = set2[10] + 1
            elif length < base * 7:
                set2[7] = set2[7] + 1
                set2[8] = set2[8] + 1
                set2[9] = set2[9] + 1
                set2[10] = set2[10] + 1
            elif length < base * 8:
                set2[8] = set2[8] + 1
                set2[9] = set2[9] + 1
                set2[10] = set2[10] + 1
            elif length < base * 9:
                set2[9] = set2[9] + 1
                set2[10] = set2[10] + 1
            elif length < base * 10:
                set2[10] = set2[10] + 1
    end = timeit.default_timer()
    with open("hopplot/caGrQc.txt", 'a') as file:
        string = "ADR--- compression:" + str(compress) + " time:" + str(end - start) + '\n'
        file.write(string)

    start = timeit.default_timer()
    edgename = "Data/min" + commontext + "-edges-ABM-" + str(compress) + ".txt"
    Gs3 = nx.read_edgelist(edgename)
    paths3 = dict(nx.shortest_path_length(Gs3))
    set3 = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    num3 = 0
    for source in paths3:
        num3 += (len(paths3[source]) - 1)
        for target in paths3[source]:
            length = paths3[source][target] *(compress+0.65*(1-compress))
            if length < base:
                set3[1] = set3[1] + 1
                set3[2] = set3[2] + 1
                set3[3] = set3[3] + 1
                set3[4] = set3[4] + 1
                set3[5] = set3[5] + 1
                set3[6] = set3[6] + 1
                set3[7] = set3[7] + 1
                set3[8] = set3[8] + 1
                set3[9] = set3[9] + 1
                set3[10] = set3[10] + 1
            elif length < base * 2:
                set3[2] = set3[2] + 1
                set3[3] = set3[3] + 1
                set3[4] = set3[4] + 1
                set3[5] = set3[5] + 1
                set3[6] = set3[6] + 1
                set3[7] = set3[7] + 1
                set3[8] = set3[8] + 1
                set3[9] = set3[9] + 1
                set3[10] = set3[10] + 1
            elif length < base * 3:
                set3[3] = set3[3] + 1
                set3[4] = set3[4] + 1
                set3[5] = set3[5] + 1
                set3[6] = set3[6] + 1
                set3[7] = set3[7] + 1
                set3[8] = set3[8] + 1
                set3[9] = set3[9] + 1
                set3[10] = set3[10] + 1
            elif length < base * 4:
                set3[4] = set3[4] + 1
                set3[5] = set3[5] + 1
                set3[6] = set3[6] + 1
                set3[7] = set3[7] + 1
                set3[8] = set3[8] + 1
                set3[9] = set3[9] + 1
                set3[10] = set3[10] + 1
            elif length < base * 5:
                set3[5] = set3[5] + 1
                set3[6] = set3[6] + 1
                set3[7] = set3[7] + 1
                set3[8] = set3[8] + 1
                set3[9] = set3[9] + 1
                set3[10] = set3[10] + 1
            elif length < base * 6:
                set3[6] = set3[6] + 1
                set3[7] = set3[7] + 1
                set3[8] = set3[8] + 1
                set3[9] = set3[9] + 1
                set3[10] = set3[10] + 1
            elif length < base * 7:
                set3[7] = set3[7] + 1
                set3[8] = set3[8] + 1
                set3[9] = set3[9] + 1
                set3[10] = set3[10] + 1
            elif length < base * 8:
                set3[8] = set3[8] + 1
                set3[9] = set3[9] + 1
                set3[10] = set3[10] + 1
            elif length < base * 9:
                set3[9] = set3[9] + 1
                set3[10] = set3[10] + 1
            elif length < base * 10:
                set3[10] = set3[10] + 1
    end = timeit.default_timer()
    with open("hopplot/caGrQc.txt", 'a') as file:
        string = "ABM--- compression:" + str(compress) + " time:" + str(end - start) + '\n'
        file.write(string)

    x = [0, base, base * 2, base * 3, base * 4, base * 5, base * 6, base * 7, base * 8, base * 9, base * 10]
    y = [z / num for z in set]
    x1 = [0, base, base * 2, base * 3, base * 4, base * 5, base * 6, base * 7, base * 8, base * 9, base * 10]
    y1 = [z / num1 for z in set1]
    x2 = [0, base, base * 2, base * 3, base * 4, base * 5, base * 6, base * 7, base * 8, base * 9, base * 10]
    y2 = [z / num2 for z in set2]
    x3 = [0, base, base * 2, base * 3, base * 4, base * 5, base * 6, base * 7, base * 8, base * 9, base * 10]
    y3 = [z / num3 for z in set3]

    # save
    A = np.array([x, y])
    A1 = np.array([x1, y1])
    A2 = np.array([x2, y2])
    A3 = np.array([x3, y3])

    data = pd.DataFrame(A)
    data1 = pd.DataFrame(A1)
    data2 = pd.DataFrame(A2)
    data3 = pd.DataFrame(A3)

    book = load_workbook('Picture/hp.xlsx')
    writer = pd.ExcelWriter('Picture/hp.xlsx', engine='openpyxl')
    writer.book = book

    pagename = "caGrQc" + str(compress)
    data.to_excel(writer, sheet_name=pagename, float_format='%.5f')
    data1.to_excel(writer, sheet_name=pagename, startrow=3, float_format='%.5f')
    data2.to_excel(writer, sheet_name=pagename, startrow=6, float_format='%.5f')
    data3.to_excel(writer, sheet_name=pagename, startrow=9, float_format='%.5f')
    writer.save()

    writer.close()

    plt.plot(x, y, marker='o', mec='k', mfc='w', c='k', label=u'initial')
    plt.plot(x1, y1, marker='v', mec='#EEB422', mfc='w', c='#EEB422', label=u'UDS')
    plt.plot(x2, y2, marker='s', mec='r', mfc='w', c='r', label=u'ADR')
    plt.plot(x3, y3, marker='+', mec='#00BFFF', mfc='w', c='#00BFFF', label=u'ABM')

    plt.legend()  # 让图例生效
    plt.xticks(x, [0, base, base * 2, base * 3, base * 4, base * 5, base * 6, base * 7, base * 8, base * 9, base * 10])
    plt.margins(0)
    plt.subplots_adjust(bottom=0.15)
    plt.xlabel(u"distance")  # X
    plt.ylabel("% of vertex pairs")  # Y
    plt.title("ca-GrQc")  # title

    plt.show()


if __name__ == '__main__':
    topk()
    # link()
    # VertexDegree()
    # ShortPath()
    # centrality()
    # coefficient()
    # hopplot()
